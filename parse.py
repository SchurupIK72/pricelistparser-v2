import pandas as pd
from rapidfuzz import process, fuzz
import re
import os

DEFAULT_MIN_MATCH_SCORE = 65  # минимальный процент совпадения по умолчанию
FILTER_NUMERIC_NOISE = True   # включить фильтрацию числового шума (цены, количества)
AUTO_FIX_MOJIBAKE = True      # попытаться починить mojibake cp1251->latin1
DIAG_COLLECT_UNMATCHED = True # собирать диагностическую информацию по непросопоставленным строкам
# Брендовые / служебные слова, которые не должны побеждать в выборе артикула, если есть более "артикульные" токены
STOPWORD_BRANDS = {"СПЕЦМАШ", "СПЕЦМAШ", "СПЕЦ", "ЕВРО", "ЕВРО4"}
PURE_NUMERIC_MIN_KEEP = 4  # Мин. длина чисто цифрового токена, чтобы не быть отброшенным как шум (если не из article-колонки)
ORDER_LINE_COLUMN = "Номер строки заказа"  # последовательный номер строки исходного заказа
OUTPUT_COLUMNS = [
    # 1. Номер строки из заказа
    ORDER_LINE_COLUMN,
    # 2. Исходный текст из заказа
    "Исходные тексты",
    # 3. Артикул из номенклатуры
    "Совпадение (из номенклатуры)",
    # 4. Название из номенклатуры
    "Название (из номенклатуры)",
    # 5. Количество из заказа
    "Количество (из заказа)",
    # 6. Цена (из номенклатуры если есть, иначе клиентская)
    "Цена",
    # Далее остальные (сохраняем для логики)
    "Процент совпадения",
    "Нормализованный артикул совпадения",
    "Извлеченный артикул",
    "Нормализованный артикул клиента",
]

# Колонка для фиксации исходного (сырого) процента совпадения до возможного повышения для вариантов
RAW_SCORE_COLUMN = "Сырой процент совпадения"
# Перемещаем сырой процент в самый конец (после основных столбцов)
if RAW_SCORE_COLUMN in OUTPUT_COLUMNS:
    OUTPUT_COLUMNS = [c for c in OUTPUT_COLUMNS if c != RAW_SCORE_COLUMN]
OUTPUT_COLUMNS.append(RAW_SCORE_COLUMN)

# --- КОНФИГУРАЦИЯ КЛЮЧЕВЫХ СЛОВ ДЛЯ ФИЛЬТРА ПО СМЫСЛУ ---
# Включить защиту по ключевым словам: строгие совпадения (tier 3 со score 100) будут понижены,
# если название номенклатуры не содержит ни одного значимого ключевого слова из строки заказа.
ENABLE_KEYWORD_GUARD = True
# Минимальная длина слова, чтобы считаться ключевым
KEYWORD_MIN_KEY_LEN = 3
# Минимальное количество общих ключевых слов (>=1 по умолчанию)
KEYWORD_REQUIRED_OVERLAP = 1
# Минимальная похожесть имени (WRatio) при отсутствии ключей, чтобы НЕ понижать (fallback)
KEYWORD_ALLOW_IF_NAME_SIM = 75
# Стоп-слова (не повышающие смысл):
KEYWORD_STOPWORDS = {
    'И','В','НА','ДЛЯ','THE','A','OF','ОТ','ПО','С','К','ДО','ON','WITH','БЕЗ','КОМПЛЕКТ','НАБОР','ЗАДНИЙ','ПЕРЕДНИЙ','ЛЕВЫЙ','ПРАВЫЙ','ПР-ВО','ПАРА'
}


def smart_engine(path: str) -> str:
    ext = os.path.splitext(path.lower())[1]
    if ext == ".xlsx":
        return "openpyxl"
    if ext == ".xls":
        return "xlrd"
    return "openpyxl"


# Надстройка над pandas.read_excel для старых .xls без CODEPAGE.
# xlrd при отсутствии CODEPAGE падает в iso-8859-1 -> возникают кракозябры.
# Пробуем: обычное чтение -> если мало кириллицы, пробуем побайтно перекодировать ячейки.
def read_legacy_xls(path, **kwargs):
    df = pd.read_excel(path, **kwargs)
    # Быстрая эвристика: если файл .xls и практически нет кириллицы, попробуем перебор перекодирования.
    if os.path.splitext(path.lower())[1] == '.xls':
        sample = " ".join(df.astype(str).head(20).fillna("").values.flatten())
        cyr = len(re.findall(r"[А-Яа-я]", sample))
        # Если кириллицы <1% символов, а есть много символов 'Ð'/'Ñ', пробуем восстановить
        if cyr < 5 and re.search(r"[ÐÑÃÂ]", sample):
            def fix_cell(v):
                if not isinstance(v, str):
                    return v
                if re.search(r"[А-Яа-я]", v):
                    return v
                try:
                    b = v.encode('latin-1', errors='ignore')
                    cand = b.decode('utf-8', errors='ignore')
                    if re.search(r"[А-Яа-я]", cand):
                        return cand
                    # Вторая попытка cp1251
                    cand2 = b.decode('cp1251', errors='ignore')
                    if re.search(r"[А-Яа-я]", cand2):
                        return cand2
                except Exception:
                    return v
                return v
            for col in df.columns:
                df[col] = df[col].apply(fix_cell)
    return df


def normalize_article(article: str) -> str:
    if not isinstance(article, str):
        return ""
    article = article.upper()
    tokens = re.findall(r"[A-ZА-Я0-9]+", article)
    tokens = sorted(tokens)
    return "".join(tokens)


def get_article_core(article: str) -> str:
    if not isinstance(article, str):
        return ""
    # Нормализуем регистр и нестандартные разделители, убираем типовые хвостовые суффиксы
    article = article.upper()
    # нормализуем дефисы и NBSP, как в extract_articles
    hyphens = "\u2010\u2011\u2012\u2013\u2014\u2212"  # ‐ ‑ ‒ – — −
    trans = {ord(c): "-" for c in hyphens}
    trans[0xA0] = " "  # NBSP -> space
    article = article.translate(trans)

    # Многократно срезаем типовые текстовые хвосты (СПЕЦМАШ/РК/СБ/Р/ВN) и затем короткие числовые вариации (-01/-10 и т.п.)
    # Сначала текстовые суффиксы
    while True:
        new_article = re.sub(r"(?:-(?:СПЕЦМАШ|РК|СБ|Р|В\d+))$", "", article)
        if new_article == article:
            break
        article = new_article
    # Затем одно- до трёхзначные числовые суффиксы (в т.ч. варианты -10Р, -01А)
    # ВАЖНО: для "точечных" конструкторских кодов (есть точки внутри, например У.036.57.000-02)
    # суффикс -02 является частью базового артикула, поэтому его не срезаем.
    if "." not in article:
        while True:
            new_article = re.sub(r"-\d{1,3}[A-ZА-Я]?$", "", article)
            if new_article == article:
                break
            article = new_article

    tokens = re.findall(r"[A-ZА-Я0-9]+", article)
    tokens = sorted(tokens)
    return "".join(tokens)

def extract_numeric_core(article: str) -> str:
    """Возвращает самую длинную числовую последовательность (ядро), длиной >=5.
    Используется для приоритезации кандидатов с одинаковым базовым номером.
    """
    if not isinstance(article, str):
        return ""
    nums = re.findall(r"[0-9]+", article)
    nums = sorted(nums, key=len, reverse=True)
    for n in nums:
        if len(n) >= 5:
            return n
    return nums[0] if nums else ""


def extract_articles(text: str):
    if not isinstance(text, str):
        return []
    text = text.upper()
    # Нормализуем дефисы (– — ‑ − …) и неразрывные пробелы
    hyphens = "\u2010\u2011\u2012\u2013\u2014\u2212"  # ‐ ‑ ‒ – — −
    trans = {ord(c): "-" for c in hyphens}
    trans[0xA0] = ord(" ")  # NBSP -> space
    text = text.translate(trans)
    # Ищем группы из букв/цифр/дефисов/точек/слэшей
    # Минимум 3 символа, чтобы ловить короткие конструкции типа "6ММ".
    # Чтобы не тащить мусор (например, трёхбуквенные слова без цифр),
    # оставляем либо токены с цифрой, либо длиной >= 4.
    raw = re.findall(r"[A-ZА-Я0-9\-\./]{3,}", text)
    candidates = [t for t in raw if any(ch.isdigit() for ch in t) or len(t) >= 4]
    return candidates


# Определение триггеров для комплектов/наборов в тексте
KIT_TRIGGERS_RE = re.compile(r"\b(КОМПЛЕКТ|КОМПЛ\.?|НАБОР|НАБ\.)\b", re.IGNORECASE)

def has_kit_trigger(text: str) -> bool:
    if not isinstance(text, str):
        return False
    return KIT_TRIGGERS_RE.search(text) is not None


# Универсальная очистка значения ячейки: приводим к строке, убираем лишнее,
# чиним типичные случаи Excel, когда артикул считался числом (float с .0 / экспонента)
def _clean_cell_value(val):
    if pd.isna(val):
        return ""
    # Если float без десятичной части -> целое
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        # Замена запятой на точку для единообразия
        return str(val).replace(",", ".")
    # Если int
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    # Удаляем конечное .0 если это артикул в текстовом виде
    if re.fullmatch(r"[0-9]+\.0", s):
        return s[:-2]
    # На всякий случай защищаемся от экспоненциальной формы (1.23457E+11)
    if re.fullmatch(r"[0-9]+\.[0-9]+E\+[0-9]+", s.upper()):
        try:
            num = float(s)
            # Форматируем без экспоненты, без потери целых
            if num.is_integer():
                return str(int(num))
            return ("%.0f" % num)
        except Exception:
            return s
    return s


# Попытка восстановить кириллицу, если текст выглядит как кракозябры
# Признак: много символов из диапазона латиницы с диакритикой, совокупность 'Ã', 'Ð', 'Ñ', 'Â', 'Ò', 'Ê'
# которые часто появляются при интерпретации UTF-8 как cp1251 или наоборот.
def maybe_fix_mojibake(s: str) -> str:
    if not isinstance(s, str) or not s:
        return s
    # Быстрый фильтр: если уже есть достаточное количество кириллицы, не трогаем
    if re.search(r"[А-Яа-я]", s):
        return s
    # Считаем число латинских "подозрительных" символов
    suspect = len(re.findall(r"[ÃÂÐÑÒÕÊ]+", s))

    # Доп. эвристика для классического варианта cp1251->latin1: присутствуют символы
    # из диапазона 0xC0-0xFF (Ô, ë, à, í, µ, ö ...) но нет кириллицы и нет ASCII слов.
    if suspect == 0:
        high_range = sum(1 for ch in s if ord(ch) >= 0xC0)
        letters = sum(1 for ch in s if ch.isalpha())
        if letters and high_range / max(letters,1) >= 0.3:  # порог 30%
            suspect = high_range  # форсируем попытки восстановления
    if suspect == 0:
        return s

    candidates = []
    def score(txt: str) -> float:
        if not txt:
            return 0.0
        total = len(txt)
        cyr = len(re.findall(r"[А-Яа-я]", txt))
        # штраф за control / replacement
        bad = len(re.findall(r"[\ufffd]", txt))
        return (cyr - bad * 2) / max(total,1)

    try:
        b_latin = s.encode('latin-1', errors='ignore')
    except Exception:
        b_latin = b""

    # Стратегии:
    # 1. latin1 bytes -> utf-8
    try:
        cand1 = b_latin.decode('utf-8', errors='ignore')
        candidates.append(cand1)
    except Exception:
        pass
    # 2. latin1 bytes -> cp1251
    try:
        cand2 = b_latin.decode('cp1251', errors='ignore')
        candidates.append(cand2)
    except Exception:
        pass
    # 3. Попытка двойного раскодирования: исходная строка могла получиться из UTF-8 -> cp1251 -> latin1
    try:
        # трактуем исходную строку как cp1251 уже раскодированную посимвольно из utf-8
        b_cp = s.encode('cp1251', errors='ignore')
        cand3 = b_cp.decode('utf-8', errors='ignore')
        candidates.append(cand3)
    except Exception:
        pass
    # 4. Попробовать повторно пропустить cand1 через тот же цикл (иногда два слоя)
    more = []
    for c in list(candidates):
        if c and not re.search(r"[А-Яа-я]", c):
            try:
                more.append(c.encode('latin-1', errors='ignore').decode('utf-8', errors='ignore'))
            except Exception:
                pass
    candidates.extend(more)

    # Оцениваем и выбираем лучшего
    best = s
    best_score = -1
    for c in candidates:
        sc = score(c)
        if sc > best_score and re.search(r"[А-Яа-я]", c):
            best_score = sc
            best = c
    # Требуем хотя бы немного кириллицы
    if best is not s and best_score > 0:
        return best
    return s


def find_header_row(path: str, sheet_name=0, search_terms=("Артикул", "Код", "Номер", "Товар")) -> int:
    preview = read_legacy_xls(
        path, sheet_name=sheet_name, header=None, nrows=60, engine=smart_engine(path)
    )
    for i in range(len(preview)):
        row_values = (
            preview.iloc[i]
            .astype(str)
            .str.replace("\n", " ")
            .str.strip()
            .str.lower()
            .tolist()
        )
        for term in search_terms:
            t = term.lower()
            if any(t == v or v.startswith(t) for v in row_values):
                return i
    return -1


def find_header_row_strict(
    path: str,
    sheet_name=0,
    search_terms=("Номенклатура", "Артикул", "Цена"),
    min_matches: int = 2,
    preview_rows: int = 80,
):
    preview = read_legacy_xls(
        path, sheet_name=sheet_name, header=None, nrows=preview_rows, engine=smart_engine(path)
    )
    terms_lower = [t.lower() for t in search_terms]
    for i in range(len(preview)):
        row_values = (
            preview.iloc[i]
            .astype(str)
            .str.replace("\n", " ")
            .str.strip()
            .str.lower()
            .tolist()
        )
        exact_count = sum(1 for t in terms_lower if t in row_values)
        if exact_count >= min_matches:
            return i
    return -1


def promote_header_if_found(df: pd.DataFrame, search_terms=("Артикул", "Код", "Номер", "Товар")) -> pd.DataFrame:
    """Если DataFrame прочитан без заголовков, попробуем найти строку-заголовок внутри первых ~60 строк
    и поднять её в columns. Возвращает обновлённый DataFrame (или исходный, если не найдено).
    """
    try:
        max_rows = min(len(df), 60)
        terms = [t.lower() for t in search_terms]
        for i in range(max_rows):
            row_values = (
                df.iloc[i]
                .astype(str)
                .str.replace("\n", " ")
                .str.strip()
                .str.lower()
                .tolist()
            )
            # ищем точное вхождение хотя бы одного из терминов
            if any(t in row_values for t in terms):
                # поднимаем строку i как заголовки
                new_cols = df.iloc[i].astype(str).str.strip().tolist()
                df2 = df.iloc[i + 1 :].copy()
                df2.columns = new_cols
                return df2
    except Exception:
        pass
    return df


def main_process(
    client_path: str = None,
    nom_path: str = None,
    output_path: str = None,
    min_score: int = DEFAULT_MIN_MATCH_SCORE,
    interactive: bool = False,
):
    if interactive:
        print("=== Интерактивный режим ===")
        client_path = (
            input(f"Файл клиента [{client_path or 'price_client.xlsx'}]: ").strip()
            or (client_path or "price_client.xlsx")
        )
        nom_path = (
            input(f"Файл номенклатуры [{nom_path or 'nomenclature.xlsx'}]: ").strip()
            or (nom_path or "nomenclature.xlsx")
        )
        output_path = (
            input(f"Файл результата [{output_path or 'result.xlsx'}]: ").strip()
            or (output_path or "result.xlsx")
        )
        min_score_input = input(
            f"Минимальный процент совпадения [{min_score}]: "
        ).strip()
        try:
            min_score = int(min_score_input) if min_score_input else min_score
        except:
            pass
        print("\nПроверьте параметры:")
        print(f"  Клиентский файл: {client_path}")
        print(f"  Номенклатура:    {nom_path}")
        print(f"  Результат:       {output_path}")
        print(f"  Порог:           {min_score}")
        ok = input("Продолжить? (Y/n): ").strip().lower()
        if ok == "n":
            print("Отменено пользователем.")
            return 0

    client_path = client_path or "price_client.xlsx"
    nom_path = nom_path or "nomenclature.xlsx"
    user_provided_output = bool(output_path)
    # Не присваиваем дефолт сразу, чтобы уметь сгенерировать динамическое имя
    if not output_path:
        output_path = None

    if not os.path.exists(client_path):
        raise FileNotFoundError(f"Файл клиента не найден: {client_path}")
    if not os.path.exists(nom_path):
        raise FileNotFoundError(f"Файл номенклатуры не найден: {nom_path}")

    print(f"[INFO] Файл клиента: {client_path}")
    print(f"[INFO] Файл номенклатуры: {nom_path}")
    # Лог пока отложим до генерации итогового пути
    print(f"[INFO] Порог совпадения: {min_score}")

    client_header = find_header_row(client_path)
    client_df = read_legacy_xls(
        client_path,
        engine=smart_engine(client_path),
        header=(client_header if client_header is not None and client_header >= 0 else None),
    )
    if client_header is None or client_header < 0:
        client_df = promote_header_if_found(
            client_df, search_terms=("Артикул", "Код", "Номер", "Товар")
        )

    # Для номенклатуры ищем строгую строку заголовка, чтобы не спутать с содержимым
    nom_header = find_header_row_strict(
        nom_path, search_terms=("Номенклатура", "Артикул", "Цена"), min_matches=2
    )
    if nom_header is None or nom_header < 0:
        # fallback: мягкий поиск
        nom_header = find_header_row(nom_path, search_terms=("Артикул", "Номенклатура", "Цена"))
    nomenclature_df = read_legacy_xls(
        nom_path,
        engine=smart_engine(nom_path),
        header=(nom_header if nom_header is not None and nom_header >= 0 else None),
    )
    if nom_header is None or nom_header < 0:
        nomenclature_df = promote_header_if_found(
            nomenclature_df, search_terms=("Номенклатура", "Артикул", "Цена")
        )
    nomenclature_df.rename(columns=lambda c: str(c).strip(), inplace=True)

    if "Артикул" not in nomenclature_df.columns:
        raise RuntimeError("Не найдена колонка 'Артикул' в номенклатуре")

    # === Варианты артикулов: базовый / -СПЕЦМАШ / -PRO-СПЕЦМАШ ===
    SPEC_VARIANT_RE = re.compile(r"(?:-PRO)?-СПЕЦМАШ$", re.IGNORECASE)

    def variant_base(a: str) -> str:
        if not isinstance(a, str):
            return ""
        return SPEC_VARIANT_RE.sub("", a.upper()).strip()

    def variant_rank(a: str) -> int:
        au = a.upper()
        if au.endswith("-PRO-СПЕЦМАШ"):
            return 0  # сначала PRO
        if au.endswith("-СПЕЦМАШ"):
            return 1  # потом просто -СПЕЦМАШ
        return 2      # затем базовый

    base_to_variants = {}
    for idx_v, a_v in enumerate(nomenclature_df["Артикул"].astype(str)):
        base_key = variant_base(a_v)
        base_to_variants.setdefault(base_key, set()).add(idx_v)

    nomenclature_df["Нормализованный артикул"] = nomenclature_df["Артикул"].apply(normalize_article)
    nomenclature_df["Базовое ядро"] = nomenclature_df["Артикул"].apply(get_article_core)
    # Доп. признаки для сопоставления комплектов: буквенная сигнатура и набор числовых токенов
    def _letter_sig(s: str) -> str:
        if not isinstance(s, str):
            return ""
        s = s.upper()
        hyphens = "\u2010\u2011\u2012\u2013\u2014\u2212"
        trans = {ord(c): "-" for c in hyphens}
        trans[0xA0] = " "
        s = s.translate(trans)
        return "".join(sorted(re.findall(r"[A-ZА-Я]+", s)))

    def _num_tokens(s: str):
        if not isinstance(s, str):
            return []
        return re.findall(r"\d+", s.upper())

    nomenclature_df["LETTER_SIG"] = nomenclature_df["Артикул"].apply(_letter_sig)
    nomenclature_df["NUM_TOKENS"] = nomenclature_df["Артикул"].apply(_num_tokens)
    nomenclature_articles = nomenclature_df["Нормализованный артикул"].tolist()
    # Множество чисто цифровых нормализованных артикулов из номенклатуры (редко, но бывают)
    digit_only_norms = {a for a in nomenclature_articles if a.isdigit()}
    # Карта склеенных цифр (удаляем всё нецифровое) -> индекс. Нужна, чтобы сопоставлять варианты
    # вида 5340-1308110 и 53401308110 как идентичные.
    digits_collapse_to_index = {}
    for idx, art in enumerate(nomenclature_df["Артикул"].astype(str)):
        collapsed = re.sub(r"\D", "", art)
        if len(collapsed) >= 8:  # только достаточно длинные, чтобы не путать с количеством
            digits_collapse_to_index.setdefault(collapsed, idx)
    # Карта для мгновенного точного совпадения по нормализованному артикулу
    norm_to_index = {}
    for idx, val in enumerate(nomenclature_articles):
        # Если дубль, оставим первый — поведение можно расширить при необходимости
        norm_to_index.setdefault(val, idx)

    # Карта базового ядра -> индекс (первый встретившийся). Нужна для сопоставления
    # "64221-3502111" с "64221-3502111-10-СПЕЦМАШ" на 100%.
    base_core_to_index = {}
    for idx, core in enumerate(nomenclature_df["Базовое ядро"].tolist()):
        base_core_to_index.setdefault(core, idx)

    # Индекс по буквенной сигнатуре -> индексы
    letter_sig_to_indices = {}
    for idx, sig in enumerate(nomenclature_df["LETTER_SIG"].tolist()):
        letter_sig_to_indices.setdefault(sig, []).append(idx)

    client_article_cols = [
        col
        for col in client_df.columns
        if any(k in str(col).lower() for k in ["артик", "код", "номер", "номенк"])
    ]
    description_cols = [
        c
        for c in client_df.columns
        if any(k in str(c).lower() for k in ["товар", "опис", "наимен", "назв"])
    ]
    quantity_cols = [
        col
        for col in client_df.columns
        if any(k in str(col).lower() for k in ["кол-во", "количество", "qty", "шт", "заказ"])
    ]
    price_cols = [
        col
        for col in client_df.columns
        if any(k in str(col).lower() for k in ["цена", "стоим", "price", "руб", "cost", "amount"])
    ]

    # Если не нашли ни одного столбца артикула/описания — делаем мягкий фолбэк:
    # считаем, что ВСЕ текстовые столбцы могут содержать полезные данные.
    if not client_article_cols and not description_cols:
        print("[WARN] Не удалось распознать заголовки с артикулами/описанием. Включён резервный режим: все столбцы будут участвовать в разборе.")
        description_cols = client_df.columns.tolist()

    # Попробуем найти название товара и цену в номенклатуре
    nom_name_col = None
    nom_price_col = None
    for c in nomenclature_df.columns:
        lc = str(c).lower()
        if nom_name_col is None and any(k in lc for k in ["номенк", "наимен", "назв", "товар", "опис"]):
            nom_name_col = c
        if nom_price_col is None and any(k in lc for k in ["цена", "стоим", "price", "руб", "cost", "amount"]):
            nom_price_col = c
        if nom_name_col and nom_price_col:
            break

    # Индекс по числовому ядру артикула -> список индексов в номенклатуре
    num_core_to_indices = {}
    for idx, art in enumerate(nomenclature_df["Артикул"].astype(str)):
        nc = extract_numeric_core(art)
        if nc:
            num_core_to_indices.setdefault(nc, []).append(idx)

    results = []
    # Будем копить непросопоставленные строки для отдельной вкладки
    unmatched_rows = []  # каждый элемент: dict с сырыми текстами, извлеченными токенами, причиной и исходными колонками

    def token_quality(tok: str) -> int:
        """Эвристический приоритет "артикульности" токена.
        Более высокий => предпочтительнее при одинаковом результате сопоставления.
        Критерии:
          5: длинная (>=8) последовательность цифр +/- разделители (похоже на артикул/код)
          4: содержит >=5 цифр или смесь цифр и букв длиной >=5
          3: содержит цифры, длина >=3
          1: чисто буквы длиной >=4 (общие слова)
          0: бренд / стоп-слово
        """
        if not tok:
            return -1
        up = tok.upper()
        letters = re.sub(r"[^A-ZА-Я]", "", up)
        digits = re.sub(r"\D", "", up)
        if up in STOPWORD_BRANDS:
            return 0
        if len(digits) >= 8:
            return 5
        if (letters and digits and len(up) >= 5) or len(digits) >= 5:
            return 4
        if digits:
            return 3
        if len(letters) >= 4:
            return 1
        return 0

    for line_idx, (_, row) in enumerate(client_df.iterrows(), start=1):
        # Множество уже добавленных артикулов номенклатуры для данной строки заказа (чтобы избежать дублей)
        emitted_articles_this_row = set()
        unmatched_reason = None
        raw_texts = []
        raw_texts_info = []  # (text, source_col_lower)
        for col in client_article_cols + description_cols:
            val = row.get(col, "")
            val_clean = _clean_cell_value(val)
            if AUTO_FIX_MOJIBAKE:
                val_clean = maybe_fix_mojibake(val_clean)
            if val_clean:
                if '|' in val_clean:
                    parts = [p.strip() for p in val_clean.split('|') if p.strip()]
                    for p in parts:
                        raw_texts.append(p)
                        raw_texts_info.append((p, str(col).lower()))
                else:
                    raw_texts.append(val_clean)
                    raw_texts_info.append((val_clean, str(col).lower()))

        # Извлекаем токены с учётом источника
        token_origin = {}  # token -> True если из article-подобного столбца
        for txt, col_lower in raw_texts_info:
            tokens_here = extract_articles(txt)
            is_article_col = any(k in col_lower for k in ["артик", "код", "номер", "номенк"])
            for tok in tokens_here:
                token_origin.setdefault(tok, is_article_col)

        extracted_all = list(token_origin.keys())

        # Фильтрация очевидного числового шума: цены, количества, даты.
        if FILTER_NUMERIC_NOISE and extracted_all:
            filtered = []
            for tok in extracted_all:
                t = tok.strip()
                t_compact = t.replace(" ", "")
                # Если содержит буквы — оставляем
                if re.search(r"[A-ZА-Я]", t_compact):
                    filtered.append(t)
                    continue
                # Десятичные с точкой или запятой: отличаем цену (обычно <=2 знаков после разделителя)
                # от артикулов вида 11.8407010 (длинная вторая часть).
                if re.fullmatch(r"\d+[\.,]\d+", t_compact):
                    # Определяем длину дробной части
                    if '.' in t_compact:
                        left, right = t_compact.split('.', 1)
                    elif ',' in t_compact:
                        left, right = t_compact.split(',', 1)
                    else:
                        left, right = t_compact, ''
                    # Если дробная часть короткая (<=2) — трактуем как цену и фильтруем.
                    # Если длинная (>=3) — оставляем как потенциальный артикул с точкой.
                    if len(right) <= 2:
                        continue
                # Дата
                if re.fullmatch(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", t_compact):
                    continue
                # Короткие чисто цифровые: фильтруем если не из article-колонки
                if t_compact.isdigit() and t_compact not in digit_only_norms:
                    if len(t_compact) < PURE_NUMERIC_MIN_KEEP and not token_origin.get(tok, False):
                        continue
                filtered.append(t)
            extracted_all = filtered

        # Выбираем лучший матч по всем кандидатам, а не первый выше порога
        chosen = None  # tuple: (art, norm_art, best_row, best_score, priority_tuple)
        raw_join_upper = (" | ".join(raw_texts)).upper() if raw_texts else ""
        for art in extracted_all:
            norm_art = normalize_article(art)
            client_core = get_article_core(art)
            best_row = None
            best_score = -1
            client_sig = "".join(sorted(re.findall(r"[A-ZА-Я]+", art.upper())))
            client_nums = re.findall(r"\d+", art.upper())
            q = token_quality(art)
            is_variant_suffix_token = art.upper().endswith("-СПЕЦМАШ") or art.upper().endswith("-PRO-СПЕЦМАШ")

            # Подготовка ключевых слов строки (делаем один раз на первую итерацию — кешируем в локальной переменной)
            if 'row_keywords' not in locals():
                joined_lower = (" ".join(raw_texts)).lower()
                cand_words = re.findall(r"[a-zа-я0-9_-]{%d,}" % KEYWORD_MIN_KEY_LEN, joined_lower, flags=re.IGNORECASE)
                row_keywords = {w for w in cand_words if w.upper() not in STOPWORD_BRANDS and w.upper() not in KEYWORD_STOPWORDS}

            # 0) Быстрое точное совпадение по склеенным цифрам (форматы с/без дефисов)
            if best_row is None:
                digits_collapsed = re.sub(r"\D", "", art)
                if len(digits_collapsed) >= 8:
                    idx_digits = digits_collapse_to_index.get(digits_collapsed)
                    if idx_digits is not None:
                        best_row = nomenclature_df.iloc[idx_digits]
                        best_score = 100
                        priority = (q, 3, 100, len(norm_art))  # (quality, tier, score, length)

            # 1) Точное совпадение по нормализованному артикулу
            if best_row is None:
                exact_idx = norm_to_index.get(norm_art)
                if exact_idx is not None:
                    best_row, best_score = nomenclature_df.iloc[exact_idx], 100
                    priority = (q, 3, 100, len(norm_art))  # tier 3 highest now
            if best_row is None:
                # 1b) Точное совпадение по базовому ядру (без вариантных суффиксов)
                if client_core:
                    base_idx = base_core_to_index.get(client_core)
                    if base_idx is not None:
                        best_row, best_score = nomenclature_df.iloc[base_idx], 100
                        priority = (q, 3, 100, len(norm_art))
                    else:
                        priority = (q, 0, 0, 0)
                else:
                    priority = (q, 0, 0, 0)

                # 1c) Расширенный вариант/комплект: числовые токены клиента ⊆ токенов номенклатуры.
                # Кандидатов берём по самому длинному числовому ядру (ускорение и точность семейства).
                if best_row is None and client_nums:
                    cand_best = None
                    cand_best_score = -1
                    cand_best_extra = -1
                    nc_long = extract_numeric_core(art)
                    candidate_idxs = num_core_to_indices.get(nc_long, []) if nc_long else []
                    for idx in candidate_idxs:
                        nom_row = nomenclature_df.iloc[idx]
                        nom_nums_list = nom_row["NUM_TOKENS"] if isinstance(nom_row.get("NUM_TOKENS"), list) else []
                        nom_nums = set(nom_nums_list)
                        if set(client_nums).issubset(nom_nums) and len(nom_nums) >= len(client_nums):
                            extra = len(nom_nums) - len(client_nums)
                            base = 95 - max(0, extra - 1) * 5
                            score_here = 100 if has_kit_trigger(raw_join_upper) else base
                            # предпочитаем более "полные" комплекты (больше extra) при равном score
                            if (score_here > cand_best_score) or (score_here == cand_best_score and extra > cand_best_extra):
                                cand_best = nom_row
                                cand_best_score = score_here
                                cand_best_extra = extra
                    if cand_best is not None:
                        best_row = cand_best
                        best_score = cand_best_score
                        priority = (q, 2 if best_score >= 100 else 1, best_score, len(norm_art))
                # 2) Приоритет по одинаковому числовому ядру + сравнение по названию
                if best_row is None:  # не сработали точные
                    nc = extract_numeric_core(art)
                    if nc and nc in num_core_to_indices:
                        name_best = -1
                        name_best_row = None
                        for idx in num_core_to_indices[nc]:
                            nom_row = nomenclature_df.iloc[idx]
                            name_val = str(nom_row.get(nom_name_col, "")) if nom_name_col else ""
                            name_score = fuzz.WRatio(raw_join_upper, name_val.upper()) if name_val else 0
                            if name_score > name_best:
                                name_best = name_score
                                name_best_row = nom_row
                        if name_best_row is not None:
                            best_row = name_best_row
                            best_score = name_best  # используем как общую уверенность
                            priority = (q, 1, name_best, len(norm_art))  # tier 1 mid

                # 3) Общий fuzzy-поиск по нормализованным артикулам (если ещё не выбрали)
                if best_row is None:
                    for match, score, idx in process.extract(
                        norm_art, nomenclature_articles, scorer=fuzz.WRatio, limit=10
                    ):
                        nom_row = nomenclature_df.iloc[idx]
                        nomen_core = get_article_core(nom_row["Артикул"])
                        if client_core and nomen_core == client_core:
                            best_row, best_score = nom_row, 100
                            priority = (q, 3, 100, len(norm_art))
                            break
                        elif score > best_score:
                            best_row, best_score = nom_row, score
                            priority = (q, 0, score, len(norm_art))

            if best_row is not None:
                # Дополнительное ослабление при низкой релевантности названия: если совпадение 100% ("жёсткая" ветка)
                # но текст строки не похож на название номенклатуры (например, артикул турбокомпрессора в строке про крыло),
                # понижаем tier, чтобы другой кандидат мог победить.
                try:
                    name_val = str(best_row.get(nom_name_col, "")) if nom_name_col else ""
                except Exception:
                    name_val = ""
                if name_val and best_score == 100 and priority[1] == 3:
                    # similarity по объединённому тексту строки
                    try:
                        name_sim = fuzz.WRatio(raw_join_upper, name_val.upper()) if raw_join_upper else 0
                    except Exception:
                        name_sim = 0
                    # Совпадение базового ядра клиента и найденного артикула
                    base_match_for_exact = False
                    if client_core and isinstance(best_row.get("Артикул"), str):
                        base_match_for_exact = (client_core == get_article_core(best_row["Артикул"]))
                    degrade = False
                    # 1) Старое правило: низкая похожесть и нет совпадения базового ядра
                    if name_sim < 60 and not base_match_for_exact:
                        degrade = True
                    # 2) Ключевые слова: если включено, требуем пересечение ключей (кроме случая когда высокая похожесть имени)
                    if ENABLE_KEYWORD_GUARD:
                        # Ключевые слова из названия номенклатуры
                        name_words = set(re.findall(r"[a-zа-я0-9_-]{%d,}" % KEYWORD_MIN_KEY_LEN, name_val.lower(), flags=re.IGNORECASE))
                        name_keywords = {w for w in name_words if w.upper() not in STOPWORD_BRANDS and w.upper() not in KEYWORD_STOPWORDS}
                        overlap = row_keywords & name_keywords if row_keywords else set()
                        if not overlap and name_sim < KEYWORD_ALLOW_IF_NAME_SIM:
                            degrade = True
                    if degrade:
                        priority = (priority[0], 1, priority[2], priority[3])
                if chosen is None:
                    chosen = (art, norm_art, best_row, best_score, priority)
                else:
                    # сравниваем по priority tuple (качество -> tier -> score -> длина нормализованного)
                    if priority > chosen[4]:
                        chosen = (art, norm_art, best_row, best_score, priority)

            # (дубликат старой логики удален)

        matched = False
        if chosen is not None and chosen[3] >= min_score:
            art, norm_art, best_row, best_score, _ = chosen
            # Цена и количество из заказа (первое найденное поле)
            price_val_client = None
            for pc in price_cols:
                v = row.get(pc, None)
                if pd.notna(v) and str(v).strip() != "":
                    price_val_client = v
                    break
            qty_val = None
            for qc in quantity_cols:
                v = row.get(qc, None)
                if pd.notna(v) and str(v).strip() != "":
                    qty_val = v
                    break

            # Строгое условие расширения вариантов: либо точные 100%, либо суффикс варианта с высоким скором и совпадением базового ядра
            art_name_upper = str(best_row["Артикул"]).upper()
            client_core_now = get_article_core(art)
            best_core_now = get_article_core(str(best_row["Артикул"]))
            base_match = client_core_now and best_core_now and client_core_now == best_core_now
            has_variant_suffix = art_name_upper.endswith("-СПЕЦМАШ") or art_name_upper.endswith("-PRO-СПЕЦМАШ")
            expand_variants = (
                best_score == 100 or (
                    has_variant_suffix and best_score >= 95 and base_match
                )
            )

            if expand_variants:
                original_score = best_score
                base_key = variant_base(str(best_row["Артикул"]))
                variant_indices = list(base_to_variants.get(base_key, {best_row.name}))
                variant_indices.sort(key=lambda i: variant_rank(str(nomenclature_df.iloc[i]["Артикул"])))
                for vidx in variant_indices:
                    vrow = nomenclature_df.iloc[vidx]
                    art_nom = str(vrow["Артикул"]) if isinstance(vrow.get("Артикул"), str) else str(vrow.get("Артикул", ""))
                    # Пропускаем, если уже выводили этот артикул для текущей строки заказа
                    if art_nom in emitted_articles_this_row:
                        continue
                    emitted_articles_this_row.add(art_nom)
                    price_val_nom = vrow.get(nom_price_col, None) if nom_price_col else None
                    results.append(
                        {
                            ORDER_LINE_COLUMN: line_idx,
                            "Исходные тексты": " | ".join(raw_texts) if raw_texts else "",
                            "Извлеченный артикул": art,
                            "Нормализованный артикул клиента": norm_art,
                            "Совпадение (из номенклатуры)": vrow["Артикул"],
                            "Название (из номенклатуры)": vrow.get(nom_name_col, "") if nom_name_col else "",
                            "Нормализованный артикул совпадения": normalize_article(vrow["Артикул"]),
                            RAW_SCORE_COLUMN: original_score,
                            "Процент совпадения": 100,
                            "Цена": (price_val_nom if (price_val_nom is not None and str(price_val_nom).strip() != "") else (price_val_client if price_val_client is not None else "")),
                            "Количество (из заказа)": qty_val if qty_val is not None else "",
                        }
                    )
            else:
                price_val_nom = best_row.get(nom_price_col, None) if nom_price_col else None
                single_art_nom = str(best_row["Артикул"]) if isinstance(best_row.get("Артикул"), str) else str(best_row.get("Артикул", ""))
                if single_art_nom in emitted_articles_this_row:
                    # Уже добавлен этим или предыдущим токеном — пропускаем дублирующую запись
                    continue
                emitted_articles_this_row.add(single_art_nom)
                results.append(
                    {
                        ORDER_LINE_COLUMN: line_idx,
                        "Исходные тексты": " | ".join(raw_texts) if raw_texts else "",
                        "Извлеченный артикул": art,
                        "Нормализованный артикул клиента": norm_art,
                        "Совпадение (из номенклатуры)": best_row["Артикул"],
                        "Название (из номенклатуры)": best_row.get(nom_name_col, "") if nom_name_col else "",
                        "Нормализованный артикул совпадения": normalize_article(best_row["Артикул"]),
                        RAW_SCORE_COLUMN: best_score,
                        "Процент совпадения": best_score,
                        "Цена": (price_val_nom if (price_val_nom is not None and str(price_val_nom).strip() != "") else (price_val_client if price_val_client is not None else "")),
                        "Количество (из заказа)": qty_val if qty_val is not None else "",
                    }
                )
            matched = True

        if not matched and raw_texts:
            # Фолбэк: используем объединенный текст, чтобы не терять артикули внутри длинных строк
            txt = " | ".join(raw_texts)
            match, score, idx = process.extractOne(
                txt.upper(),
                nomenclature_df["Артикул"].astype(str).tolist(),
                scorer=fuzz.WRatio,
            )
            if score >= min_score:
                best_row = nomenclature_df.iloc[idx]

                # Цена и количество из заказа (первое найденное поле)
                price_val_client = None
                for pc in price_cols:
                    v = row.get(pc, None)
                    if pd.notna(v) and str(v).strip() != "":
                        price_val_client = v
                        break
                qty_val = None
                for qc in quantity_cols:
                    v = row.get(qc, None)
                    if pd.notna(v) and str(v).strip() != "":
                        qty_val = v
                        break

                # Цена из номенклатуры приоритетнее
                price_val_nom = best_row.get(nom_price_col, None) if nom_price_col else None

                fallback_art_nom = str(best_row["Артикул"]) if isinstance(best_row.get("Артикул"), str) else str(best_row.get("Артикул", ""))
                if fallback_art_nom not in emitted_articles_this_row:
                    emitted_articles_this_row.add(fallback_art_nom)
                    results.append(
                        {
                            ORDER_LINE_COLUMN: line_idx,
                            "Исходные тексты": txt,
                            "Извлеченный артикул": "",
                            "Нормализованный артикул клиента": "",
                            "Совпадение (из номенклатуры)": best_row["Артикул"],
                            "Название (из номенклатуры)": best_row.get(nom_name_col, "") if nom_name_col else "",
                            "Нормализованный артикул совпадения": normalize_article(best_row["Артикул"]),
                            RAW_SCORE_COLUMN: score,
                            "Процент совпадения": score,
                            "Цена": (price_val_nom if (price_val_nom is not None and str(price_val_nom).strip() != "") else (price_val_client if price_val_client is not None else "")),
                            "Количество (из заказа)": qty_val if qty_val is not None else "",
                        }
                    )
                matched = True
            else:
                unmatched_reason = f"fallback_wratio={score}<min_score"
        if not matched:
            if extracted_all:
                if chosen is not None:
                    unmatched_reason = f"best_score={chosen[3]}<min_score"
                else:
                    unmatched_reason = "no_candidate_after_filters"
            else:
                unmatched_reason = "no_tokens_extracted"
        if DIAG_COLLECT_UNMATCHED and not matched:
            # Сохраняем первые 50 непросопоставленных для анализа
            if 'diag_unmatched' not in locals():
                diag_unmatched = []
            if len(diag_unmatched) < 50:
                diag_unmatched.append({
                    'raw': raw_texts,
                    'extracted': extracted_all,
                    'reason': unmatched_reason
                })
        # Накапливаем для итоговой таблицы непросопоставленных
        if not matched:
            row_record = {
                ORDER_LINE_COLUMN: line_idx,
                'Исходные тексты': " | ".join(raw_texts) if raw_texts else "",
                'Извлеченные токены': ", ".join(extracted_all) if extracted_all else "",
                'Причина': unmatched_reason or "",
            }
            # Добавим исходные значения клиентских колонок (чтобы можно было анализировать)
            for col in client_df.columns:
                try:
                    val = row.get(col, "")
                except Exception:
                    val = ""
                row_record[f'CLIENT::{col}'] = val
            unmatched_rows.append(row_record)

    # Сформируем результирующий DataFrame
    df_result = pd.DataFrame(results)

    # Гарантируем одинаковые столбцы и порядок в любом режиме (GUI/консоль)
    for col in OUTPUT_COLUMNS:
        if col not in df_result.columns:
            df_result[col] = ""
    # Отбрасываем лишние столбцы и упорядочиваем по эталону
    df_result = df_result[OUTPUT_COLUMNS]

    # Подготовим DataFrame непросопоставленных
    df_unmatched = pd.DataFrame(unmatched_rows)

    # Если путь не был задан пользователем — формируем имя вида
    # result+<заданный_порог>.xlsx (порог = min_score) в папке клиента
    if not user_provided_output:
        client_dir = os.path.dirname(os.path.abspath(client_path)) or os.getcwd()
        client_base = os.path.splitext(os.path.basename(client_path))[0]
        # Формат: result+<порог>+<имя_исходного_файла_без_расширения>.xlsx
        dynamic_name = f"result+{min_score}+{client_base}.xlsx"
        output_path = os.path.join(client_dir, dynamic_name)
    print(f"[INFO] Файл результата: {output_path}")

    # Запишем в Excel и подсветим зеленым строки со 100% совпадением
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sheet_name = "Sheet1"
        df_result.to_excel(writer, index=False, sheet_name=sheet_name)
        # Имя вкладки для непросопоставленных: OutResult+<название результирующей вкладки>
        unmatched_sheet_name = f"OutResult+{sheet_name}"[:31]  # Excel limit 31 chars
        if not df_unmatched.empty:
            df_unmatched.to_excel(writer, index=False, sheet_name=unmatched_sheet_name)

        # Получаем лист и применяем заливку строкам с 100%
        ws = writer.sheets[sheet_name]
        try:
            from openpyxl.styles import PatternFill
        except Exception:
            PatternFill = None

        # Закрепляем шапку (freeze header) для удобства просмотра больших таблиц
        try:
            ws.freeze_panes = ws["A2"]  # фиксируем первую строку
            if not df_unmatched.empty:
                ws_unmatched = writer.sheets.get(unmatched_sheet_name)
                if ws_unmatched is not None:
                    ws_unmatched.freeze_panes = ws_unmatched["A2"]
        except Exception:
            pass

        if PatternFill is not None and "Процент совпадения" in df_result.columns:
            green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
            dark_green_fill = PatternFill(fill_type="solid", start_color="82B366", end_color="82B366")  # более тёмный для семейств
            variant_suffixes = ("-СПЕЦМАШ", "-PRO-СПЕЦМАШ")
            # Подсчёт количества строк на каждый номер заказа
            try:
                line_counts = df_result[ORDER_LINE_COLUMN].value_counts()
                multi_line_numbers = {ln for ln, cnt in line_counts.items() if cnt >= 2}
            except Exception:
                multi_line_numbers = set()
            # Предполагаем доступ к функции variant_base (определена выше в main_process)
            # Построим карту: номер строки -> список артикулов совпадения
            line_to_articles = {}
            if ORDER_LINE_COLUMN in df_result.columns and "Совпадение (из номенклатуры)" in df_result.columns:
                for ln, art_match in zip(df_result[ORDER_LINE_COLUMN], df_result["Совпадение (из номенклатуры)"]):
                    try:
                        line_to_articles.setdefault(ln, []).append(str(art_match))
                    except Exception:
                        pass
            # Определим для каких номеров строк есть семейства с вариантами (наличие хотя бы одного вариантного суффикса)
            family_lines_with_variants = set()
            line_base_keys_with_variant = {}
            for ln, arts in line_to_articles.items():
                if ln not in multi_line_numbers:
                    continue
                has_variant = any(isinstance(a, str) and a.upper().endswith(variant_suffixes) for a in arts)
                if not has_variant:
                    continue
                family_lines_with_variants.add(ln)
                # Собираем базовые ключи для всех артикулов в этой строке
                base_keys = set()
                for a in arts:
                    if isinstance(a, str):
                        try:
                            base_keys.add(variant_base(a))
                        except Exception:
                            pass
                line_base_keys_with_variant[ln] = base_keys
            # Индексы нужных столбцов
            try:
                match_col_index = OUTPUT_COLUMNS.index("Совпадение (из номенклатуры)") + 1
            except ValueError:
                match_col_index = None
            try:
                order_line_col_index = OUTPUT_COLUMNS.index(ORDER_LINE_COLUMN) + 1
            except ValueError:
                order_line_col_index = None
            # Цикл по строкам листа (данные начинаются со 2-й)
            for r_idx, score in enumerate(df_result["Процент совпадения"], start=2):
                is_hundred = False
                try:
                    val = float(str(score).replace(",", "."))
                    is_hundred = abs(val - 100.0) < 1e-9
                except Exception:
                    pass
                if not is_hundred:
                    continue
                row_fill = green_fill  # default
                try:
                    line_number_val = ws.cell(row=r_idx, column=order_line_col_index).value if order_line_col_index else None
                except Exception:
                    line_number_val = None
                if match_col_index is not None and line_number_val in family_lines_with_variants:
                    try:
                        art_match_val = ws.cell(row=r_idx, column=match_col_index).value
                    except Exception:
                        art_match_val = None
                    make_dark = False
                    if isinstance(art_match_val, str):
                        up = art_match_val.upper()
                        if up.endswith(variant_suffixes):
                            make_dark = True  # сам вариант
                        else:
                            # Базовый артикул: красим если его базовый ключ совпадает с базовым ключом любого вариантного
                            try:
                                base_key = variant_base(art_match_val)
                                if base_key and base_key in line_base_keys_with_variant.get(line_number_val, set()):
                                    make_dark = True
                            except Exception:
                                pass
                    if make_dark:
                        row_fill = dark_green_fill
                for c_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = row_fill

    print(f"✅ Готово! Найдено совпадений: {len(results)}")
    if DIAG_COLLECT_UNMATCHED and 'diag_unmatched' in locals() and diag_unmatched:
        print(f"[DIAG] Непросопоставленных строк: {len(diag_unmatched)} (показаны первые {len(diag_unmatched)})")
        for d in diag_unmatched[:5]:
            print("[DIAG] reason=", d['reason'], " raw=", " || ".join(d['raw']), " extracted=", d['extracted'])
    return len(results)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Сопоставление артикулов")
    parser.add_argument("--client", "-c", help="Файл клиента")
    parser.add_argument("--nomenclature", "-n", help="Файл номенклатуры")
    parser.add_argument("--output", "-o", help="Файл результата")
    parser.add_argument("--min-score", type=int, default=DEFAULT_MIN_MATCH_SCORE)
    parser.add_argument("--interactive", "-i", action="store_true")
    args = parser.parse_args()

    if args.interactive:
        main_process(interactive=True)
    else:
        main_process(
            client_path=args.client,
            nom_path=args.nomenclature,
            output_path=args.output,
            min_score=args.min_score,
        )
